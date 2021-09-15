import openpyxl
class XLSReader:
    def __init__(self, path):
        self.path=path
        self.wb_obj = openpyxl.load_workbook(self.path)

    def getCellData(self, sheetName, rowNum, colNum):
        sheet_obj =self.wb_obj[sheetName]
        cell_obj= sheet_obj.cell(row=rowNum, column=colNum)
        return cell_obj

    def getCellDataByColName(self, sheetName, rowNum, colName):
        sheet = self.wb_obj[sheetName]
        for cNum in range(1, sheet.max_column+1):
            extractedColName = sheet.cell(1, cNum).value
            if(extractedColName==colName):
                cellData = sheet.cell(rowNum, cNum).value
                if(cellData!=''):
                    return cellData
                else:
                    return ''




    # def checkEmptyCell(self, sheetName, rowNum, colNum):
    #     sheet_obj = self.wb_obj[sheetName]
    #     print("Hi in check empty cell function")
    #     cell_obj = sheet_obj.cell(rowNum, colNum).value
    #     print("Cell value is "+str(cell_obj))
    #     if(cell_obj == None):
    #         print("In if")
    #         return True
    #     else:
    #         print("In else")
    #         return False

    def totalRows(self, sheetName, testcasename):
        testStartRowIndex = 1
        colStartIndex = 1
        while not ((self.getCellData(sheetName, testStartRowIndex, colStartIndex).value) == testcasename):
            testStartRowIndex = testStartRowIndex + 1
        headerStartIndex = testStartRowIndex+1
        dataStartIndex = testStartRowIndex+2
        while not ((self.getCellData(sheetName, dataStartIndex, colStartIndex).value) == None):
            dataStartIndex = dataStartIndex + 1

        totalRows = (dataStartIndex-1) - headerStartIndex
        return totalRows

    def totalCols(self, sheetName, testcasename):
        testStartRowIndex = 1
        colStartIndex = 1
        while not ((self.getCellData(sheetName, testStartRowIndex, colStartIndex).value) == testcasename):
            testStartRowIndex = testStartRowIndex + 1
        headerStartRowIndex = testStartRowIndex + 1

        while not((self.getCellData(sheetName, headerStartRowIndex, colStartIndex).value) == None):
            colStartIndex = colStartIndex + 1

        totalCols = colStartIndex -1
        return totalCols

    def rowCount(self, sheetname):
        sheet = self.wb_obj[sheetname]
        return sheet.max_row

    def colCount(self, sheetname):
        sheet = self.wb_obj[sheetname]
        return sheet.max_column
